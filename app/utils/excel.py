"""
Utility functions for Excel report generation.
Uses openpyxl to create spreadsheets from report data.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict
from io import BytesIO

def generate_excel(data: List[Dict], sheet_name: str = "Report") -> bytes:
    """
    Generates an Excel file from a list of dictionaries.
    Returns the file as bytes suitable for StreamingResponse.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        ws.append(["No data available"])
    else:
        # Write header
        headers = list(data[0].keys())
        ws.append(headers)

        # Write rows
        for row in data:
            ws.append([row.get(h) for h in headers])

        # Auto-size columns
        for i, col in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col) + 2
            ws.column_dimensions[get_column_letter(i)].width = max_length

    # Save to bytes
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()
